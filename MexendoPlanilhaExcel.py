from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import os

"""
Código Não terminado #################################################################################################

Pode ser testado para mostrar ao professor,
mais informaçoes nas observacoes no final do codigo!

"""

#Caminho da planilha (De acordo com o caminho criado na máquina)
CaminhoPlanilha = "C:\\Users\\vinic\\PycharmProjects\\PIEC-1-Jos-Vinicius-Carla-Lauro-\\Controle de Estoque.xlsx"

Planilha = load_workbook(filename=CaminhoPlanilha) #Abre a planilha

#Abre a página da planilha
sheetAberta = Planilha['Controle de Estoque']

#Formatação dos dados que vão entrar
CorDados = {
    "font": Font(
        name="Arial",
        size=10,
        bold=True,
        color="000000"  # Preto
    ),
    "alignment": Alignment( #Alinhamento
        horizontal="center",
        vertical="center"
    ),
    "fill": PatternFill( # Fundo da célula
        fill_type="solid",
        fgColor="D9D9D9"  # cinza claro
    ),
    "border": Border( #Borda, todas saem preta por padrão.
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
}

#Função que adiciona os dados do produto
def adicionar_produto(produto, codigo, quantidade, estoque_minimo):

    # Descobre a próxima linha disponível
    linha = sheetAberta.max_row + 1

    # Monta os dados com data e hora atuais
    dados = [
        produto,
        codigo,
        quantidade,
        estoque_minimo,
        datetime.now().strftime("%d/%m/%Y"),  # Data atual
        datetime.now().strftime("%H:%M:%S")   # Hora atual
    ]

    # Insere cada valor na linha e aplica formatação
    for coluna, valor in enumerate(dados, start=1):
        celula = sheetAberta.cell(row=linha, column=coluna, value=valor)
        celula.font = CorDados["font"]
        celula.alignment = CorDados["alignment"]
        celula.border = CorDados["border"]
        celula.fill = CorDados["fill"]

    print(f"Produto '{produto}' adicionado na linha {linha}.")

#Loop para chamar a função e adicionar os dados
while True:

    #Layout teste
    print("\n=== Controle de Estoque ===")
    acao = input("Digite [A] para adicionar produto, [S] para sair: ").upper()

    if acao == "A":
        produto = input("Nome do produto: ")
        codigo = int(input("Código do produto: "))
        quantidade = int(input("Quantidade: "))
        estoque_minimo = int(input("Estoque mínimo: "))

        # Chama a função que insere os dados na planilha
        adicionar_produto(produto, codigo, quantidade, estoque_minimo)

    elif acao == "S":
        print("Saindo do sistema...")
        break
    else:
        print("Opção inválida, tente novamente.")



# Salva alterações ao final
Planilha.save(CaminhoPlanilha)
print("Planilha atualizada e salva.")

#Abre a planilha
os.startfile(CaminhoPlanilha)



'''
Adicionar rota para chamar a função e o loop
Fazer o teste para ver se o loop esta funcionando e colocando os itens nos lugares corretos!
- Adicionar tratamentos para melhor sofisticacao
- Adicionar tratamentos de erro
'''
